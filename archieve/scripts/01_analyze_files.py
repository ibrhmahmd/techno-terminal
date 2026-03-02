"""
Script 1: File Discovery & Excel Metadata Extraction
====================================================
Recursively scans the entire directory tree, opens every valid Excel file,
and extracts: sheet names, column headers, row counts, data types, and file metadata.

Outputs:
  _standardization_output/file_metadata.json   – full per-file detail
  _standardization_output/file_inventory.csv   – summary table

Usage:
  python 01_analyze_files.py
"""

import os
import re
import json
import csv
import hashlib
import unicodedata
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not found. Run: pip install openpyxl pandas xlrd")
    exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent
RAW_DIR = BASE_DIR / "data" / "raw"
OUTPUT_DIR = BASE_DIR / "data" / "processed" / "_standardization_output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}
SKIP_PREFIX = "~$"  # temp lock files
MAX_SAMPLE_ROWS = 5  # rows to sample for type detection
MAX_HEADER_SCAN_ROWS = 10  # rows to scan for header detection


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def normalize_header(raw):
    """Normalize a column header to a consistent snake_case key."""
    if raw is None:
        return None
    s = str(raw).strip()
    # Remove diacritics from Arabic/unicode
    s = unicodedata.normalize("NFKD", s)
    # Lowercase
    s = s.lower()
    # Replace common separators with underscore
    s = re.sub(r"[\s\-/\\\.]+", "_", s)
    # Remove anything not alphanumeric or underscore
    s = re.sub(r"[^\w]", "", s)
    # Collapse multiple underscores
    s = re.sub(r"_+", "_", s).strip("_")
    return s if s else None


def detect_cell_type(value):
    """Detect the semantic type of a single cell value."""
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        # Could be Excel date serial
        if isinstance(value, float) and 1 <= value <= 100000:
            return "numeric_or_date"
        return "numeric"
    if isinstance(value, (datetime, date)):
        return "date"
    s = str(value).strip()
    if re.match(r"^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$", s):
        return "date_string"
    if re.match(r"^\d{10,15}$", s):
        return "phone"
    if s.lower() in {
        "1",
        "0",
        "p",
        "a",
        "x",
        "✓",
        "✗",
        "yes",
        "no",
        "present",
        "absent",
        "حضر",
        "غاب",
        "h",
        "l",
        "late",
    }:
        return "attendance_marker"
    return "string"


def sha256_file(path: Path) -> str:
    """Compute SHA-256 hash of a file for deduplication."""
    h = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return "UNREADABLE"


def find_header_row(ws, max_scan=MAX_HEADER_SCAN_ROWS):
    """Find the row index (1-based) most likely to be the header row."""
    best_row = 1
    best_score = -1
    limit = min(max_scan, ws.max_row)
    for r in range(1, limit + 1):
        string_count = 0
        non_null_count = 0
        for c in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip():
                non_null_count += 1
                if isinstance(val, str):
                    string_count += 1
        # Score: rows with mostly strings score high
        score = string_count * 2 + non_null_count
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def extract_sheet_metadata(ws, sheet_name, file_path_str):
    """Extract metadata from a single worksheet."""
    if ws.max_row is None or ws.max_row == 0:
        return {
            "sheet_name": sheet_name,
            "is_empty": True,
            "header_row": None,
            "headers_raw": [],
            "headers_normalized": [],
            "fingerprint": [],
            "row_count": 0,
            "column_count": 0,
            "type_samples": {},
        }

    header_row_idx = find_header_row(ws)
    headers_raw = []
    headers_normalized = []

    for c in range(1, (ws.max_column or 0) + 1):
        cell_val = ws.cell(row=header_row_idx, column=c).value
        raw = str(cell_val).strip() if cell_val is not None else None
        norm = normalize_header(raw)
        headers_raw.append(raw)
        headers_normalized.append(norm)

    # Remove trailing None headers
    while headers_normalized and headers_normalized[-1] is None:
        headers_normalized.pop()
        headers_raw.pop()

    # Deduplicate (add suffix on collision)
    seen = {}
    deduped = []
    for h in headers_normalized:
        if h is None:
            deduped.append(None)
            continue
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    headers_normalized = deduped

    fingerprint = sorted(set(h for h in headers_normalized if h))

    # Sample data rows for type detection
    data_start = header_row_idx + 1
    data_end = min(data_start + MAX_SAMPLE_ROWS - 1, ws.max_row)
    row_count = max(0, ws.max_row - header_row_idx)

    type_samples = {}
    for col_idx, col_name in enumerate(headers_normalized, start=1):
        if col_name is None:
            continue
        types = []
        for r in range(data_start, data_end + 1):
            val = ws.cell(row=r, column=col_idx).value
            types.append(detect_cell_type(val))
        # Majority type
        from collections import Counter

        type_counts = Counter(t for t in types if t != "null")
        dominant = type_counts.most_common(1)[0][0] if type_counts else "null"
        type_samples[col_name] = dominant

    return {
        "sheet_name": sheet_name,
        "is_empty": row_count == 0,
        "header_row": header_row_idx,
        "headers_raw": headers_raw,
        "headers_normalized": headers_normalized,
        "fingerprint": fingerprint,
        "row_count": row_count,
        "column_count": len([h for h in headers_normalized if h]),
        "type_samples": type_samples,
    }


# ─────────────────────────────────────────────────────────────────────────────
# MAIN SCANNER
# ─────────────────────────────────────────────────────────────────────────────
def scan_all_files():
    records = []
    inventory_rows = []
    stats = {
        "total_files_found": 0,
        "excel_files": 0,
        "temp_lock_files": 0,
        "other_files": 0,
        "unreadable_files": 0,
        "total_sheets": 0,
        "total_rows": 0,
    }

    print(f"\n{'=' * 70}")
    print(f"  Excel File Discovery & Metadata Extraction")
    print(f"  Base directory: {RAW_DIR}")
    print(f"{'=' * 70}\n")

    hashes_seen = {}  # hash → first_file_path (deduplication)

    for root, dirs, files in os.walk(RAW_DIR):
        root_path = Path(root)

        for filename in sorted(files):
            stats["total_files_found"] += 1
            file_path = root_path / filename
            rel_path = str(file_path.relative_to(RAW_DIR))
            ext = file_path.suffix.lower()

            # ── Skip temp/lock files ──────────────────────────────────────────
            if filename.startswith(SKIP_PREFIX):
                stats["temp_lock_files"] += 1
                continue

            # ── Skip non-Excel files ─────────────────────────────────────────
            if ext not in EXCEL_EXTENSIONS:
                stats["other_files"] += 1
                continue

            stats["excel_files"] += 1
            file_size = file_path.stat().st_size
            file_hash = sha256_file(file_path)
            is_duplicate = file_hash in hashes_seen and file_hash != "UNREADABLE"
            if not is_duplicate:
                hashes_seen[file_hash] = rel_path

            print(f"  - Scanning: {rel_path}")

            record = {
                "file_path": rel_path,
                "filename": filename,
                "extension": ext,
                "size_bytes": file_size,
                "sha256": file_hash,
                "is_duplicate_of": hashes_seen.get(file_hash) if is_duplicate else None,
                "scanned_at": datetime.now().isoformat(),
                "sheets": [],
                "error": None,
            }

            # ── Open and inspect ─────────────────────────────────────────────
            try:
                wb = openpyxl.load_workbook(
                    str(file_path), read_only=True, data_only=True
                )
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    sheet_meta = extract_sheet_metadata(ws, sname, rel_path)
                    record["sheets"].append(sheet_meta)
                    stats["total_sheets"] += 1
                    stats["total_rows"] += sheet_meta["row_count"]
                wb.close()
            except Exception as e:
                stats["unreadable_files"] += 1
                record["error"] = str(e)
                print(f"    [WARN] Could not open: {e}")

            records.append(record)

            # ── Summary row for CSV ──────────────────────────────────────────
            total_rows = sum(s["row_count"] for s in record["sheets"])
            all_fingerprints = ";".join(
                "|".join(s["fingerprint"]) for s in record["sheets"]
            )
            inventory_rows.append(
                {
                    "relative_path": rel_path,
                    "filename": filename,
                    "extension": ext,
                    "size_bytes": file_size,
                    "sheet_count": len(record["sheets"]),
                    "sheet_names": ";".join(s["sheet_name"] for s in record["sheets"]),
                    "total_rows": total_rows,
                    "is_duplicate": is_duplicate,
                    "error": record["error"] or "",
                    "sha256": file_hash[:16] + "...",
                }
            )

    # ── Write outputs ────────────────────────────────────────────────────────
    metadata_path = OUTPUT_DIR / "file_metadata.json"
    with open(metadata_path, "w", encoding="utf-8") as f:
        json.dump(
            {"stats": stats, "files": records},
            f,
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    print(f"\n  [OK] Metadata written → {metadata_path}")

    inventory_path = OUTPUT_DIR / "file_inventory.csv"
    if inventory_rows:
        with open(inventory_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=inventory_rows[0].keys())
            writer.writeheader()
            writer.writerows(inventory_rows)
    print(f"  [OK] Inventory written → {inventory_path}")

    # ── Print summary ────────────────────────────────────────────────────────
    print(f"\n{'=' * 70}")
    print(f"  SCAN SUMMARY")
    print(f"  Total files found          : {stats['total_files_found']}")
    print(f"  Excel files scanned        : {stats['excel_files']}")
    print(f"  Temp/lock files skipped    : {stats['temp_lock_files']}")
    print(f"  Other files skipped        : {stats['other_files']}")
    print(f"  Unreadable/corrupt         : {stats['unreadable_files']}")
    print(f"  Total worksheets           : {stats['total_sheets']}")
    print(f"  Total data rows            : {stats['total_rows']}")
    print(f"{'=' * 70}\n")

    return records, stats


if __name__ == "__main__":
    scan_all_files()
