"""
Script 5: Master Schema Template Generator
==========================================
Generates standardized Excel template files for each business domain:
  - CourseAttendance_Template.xlsx
  - Competition_Template.xlsx
  - StaffAttendance_Template.xlsx
  - CRM_SAT_Template.xlsx
  - MasterSchema_UnifiedTable.xlsx  (the target flat table)

All templates include:
  - Color-coded header rows
  - Data validation dropdowns where applicable
  - Column width auto-sizing
  - Instructional notes row

Outputs:
  _standardization_output/templates/

Usage:
  python 05_generate_schema_template.py
"""

from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("[ERROR] openpyxl not found. Run: pip install openpyxl")
    exit(1)

BASE_DIR = Path(__file__).parent.parent
TEMPLATE_DIR = BASE_DIR / "data" / "processed" / "_standardization_output" / "templates"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

# ── Style helpers ─────────────────────────────────────────────────────────────
COLORS = {
    "header_blue": "1F4E79",
    "header_green": "375623",
    "header_orange": "843C0C",
    "header_purple": "3B1F75",
    "header_teal": "1F5F6B",
    "note_yellow": "FFF2CC",
    "alt_row": "EEF4FB",
    "white": "FFFFFF",
    "light_gray": "F2F2F2",
    "master_header": "0D0D0D",
}


def make_header_style(hex_color, bold=True, font_size=10):
    return {
        "fill": PatternFill(fill_type="solid", fgColor=hex_color),
        "font": Font(color="FFFFFF", bold=bold, size=font_size, name="Calibri"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
        ),
    }


def make_note_style():
    return {
        "fill": PatternFill(fill_type="solid", fgColor=COLORS["note_yellow"]),
        "font": Font(italic=True, size=9, color="7F6000", name="Calibri"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }


def apply_styles(cell, style_dict):
    for attr, val in style_dict.items():
        setattr(cell, attr, val)


def set_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def add_header_row(ws, row_num, headers, style, row_height=30):
    ws.row_dimensions[row_num].height = row_height
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        apply_styles(cell, style)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_note_row(ws, row_num, notes, style, row_height=35):
    ws.row_dimensions[row_num].height = row_height
    for col_idx, note in enumerate(notes, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=note)
        apply_styles(cell, style)


def add_sample_rows(ws, start_row, samples, alt_fill=None):
    for i, row_data in enumerate(samples):
        row_num = start_row + i
        fill = (
            PatternFill(fill_type="solid", fgColor=COLORS["alt_row"])
            if (i % 2 == 1)
            else None
        )
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill


def freeze_and_filter(ws, freeze_cell="A3"):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions


# ── Template 1: Course Attendance ─────────────────────────────────────────────
def create_course_attendance_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CourseAttendance"
    ws.sheet_view.showGridLines = True

    style = make_header_style(COLORS["header_blue"])
    note_style = make_note_style()

    headers = [
        ("Student Name", 22),
        ("Phone", 16),
        ("Course Name", 22),
        ("Level", 9),
        ("Session Day", 14),
        ("Session Time", 14),
        ("Instructor", 18),
        ("Date1\n(DD/MM/YYYY)", 14),
        ("Date2\n(DD/MM/YYYY)", 14),
        ("Date3\n(DD/MM/YYYY)", 14),
        ("Date4\n(DD/MM/YYYY)", 14),
        ("Date5\n(DD/MM/YYYY)", 14),
        ("Date6\n(DD/MM/YYYY)", 14),
        ("Date7\n(DD/MM/YYYY)", 14),
        ("Date8\n(DD/MM/YYYY)", 14),
        ("Source File", 28),
    ]

    notes = [
        "Full student name",
        "11-digit phone",
        "Exact course name",
        "1,2,3…",
        "Saturday/Friday…",
        "e.g. 10:00-12:00",
        "Instructor name",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "Auto-filled from source",
    ]

    add_header_row(ws, 1, headers, style)
    add_note_row(ws, 2, notes, note_style)

    samples = [
        [
            "Ahmed Mohamed",
            "01012345678",
            "HTML & CSS",
            1,
            "Saturday",
            "10:00-12:00",
            "Nada",
            1,
            1,
            0,
            1,
            1,
            None,
            None,
            None,
            "courses 2/Html sat 6-8.xlsx",
        ],
        [
            "Sara Ali",
            "01098765432",
            "Scratch",
            2,
            "Friday",
            "14:00-17:00",
            "Reem",
            1,
            0,
            1,
            1,
            None,
            None,
            None,
            None,
            "courses 2/SCRATCH FRI REEM 3.xlsx",
        ],
    ]
    add_sample_rows(ws, 3, samples)
    freeze_and_filter(ws, "A3")

    # Day validation
    dv_day = DataValidation(
        type="list",
        formula1='"Saturday,Sunday,Monday,Tuesday,Wednesday,Thursday,Friday"',
        showDropDown=False,
    )
    ws.add_data_validation(dv_day)
    dv_day.sqref = "E3:E1000"

    # Attendance validation
    dv_att = DataValidation(
        type="whole", operator="between", formula1="0", formula2="1"
    )
    ws.add_data_validation(dv_att)
    for col in range(8, 16):
        dv_att.sqref = f"{get_column_letter(col)}3:{get_column_letter(col)}1000"

    # Title
    ws.insert_rows(1)
    title_cell = ws.cell(
        row=1, column=1, value="COURSE ATTENDANCE REGISTER — Standardized Template v1.0"
    )
    title_cell.font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="D6E4F7")
    ws.merge_cells("A1:P1")
    ws.row_dimensions[1].height = 22

    path = TEMPLATE_DIR / "CourseAttendance_Template.xlsx"
    wb.save(str(path))
    print(f"  [OK] {path.name}")
    return path


# ── Template 2: Competition Tracking ─────────────────────────────────────────
def create_competition_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CompetitionTracking"

    style = make_header_style(COLORS["header_green"])
    note_style = make_note_style()

    headers = [
        ("Team Name", 22),
        ("Members (semicolon-separated)", 40),
        ("Competition", 20),
        ("Category", 18),
        ("Session Date", 16),
        ("Score / Points", 14),
        ("Round", 10),
        ("Rank", 10),
        ("Notes", 30),
        ("Source File", 28),
    ]
    notes = [
        "Official team name",
        "Name1; Name2; Name3",
        "FLL / Robofest / WRO",
        "Junior / Senior / Open",
        "DD/MM/YYYY",
        "Numeric",
        "1,2,3,Final…",
        "1st,2nd…",
        "Observations",
        "Auto-filled",
    ]

    add_header_row(ws, 1, headers, style)
    add_note_row(ws, 2, notes, note_style)

    samples = [
        [
            "Team Techno Alpha",
            "Omar; Youssef; Laila",
            "FLL",
            "Junior",
            "15/11/2024",
            285,
            "Round 1",
            "2nd",
            "",
            "courses 2/Old Courses/FLL/FLL November 2024.xlsx",
        ],
        [
            "Robo Stars",
            "Karim; Hana",
            "Robofest",
            "Senior",
            "10/03/2025",
            320,
            "Final",
            "1st",
            "Won engineering award",
            "Esraa/ROBOFEST SHEETS/robofest (4-5).xlsx",
        ],
    ]
    add_sample_rows(ws, 3, samples)
    freeze_and_filter(ws, "A3")

    ws.insert_rows(1)
    title = ws.cell(
        row=1, column=1, value="COMPETITION TRACKING — Standardized Template v1.0"
    )
    title.font = Font(bold=True, size=12, color="375623", name="Calibri")
    title.fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 22

    path = TEMPLATE_DIR / "Competition_Template.xlsx"
    wb.save(str(path))
    print(f"  [OK] {path.name}")
    return path


# ── Template 3: Staff Attendance ──────────────────────────────────────────────
def create_staff_attendance_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StaffAttendance"

    style = make_header_style(COLORS["header_orange"])
    note_style = make_note_style()

    headers = [
        ("Staff Name", 22),
        ("Role / Position", 20),
        ("Department", 18),
        ("Date1\n(DD/MM/YYYY)", 14),
        ("Date2\n(DD/MM/YYYY)", 14),
        ("Date3\n(DD/MM/YYYY)", 14),
        ("Date4\n(DD/MM/YYYY)", 14),
        ("Date5\n(DD/MM/YYYY)", 14),
        ("Date6\n(DD/MM/YYYY)", 14),
        ("Total Present", 14),
        ("Total Absent", 14),
        ("Attendance %", 14),
        ("Source File", 28),
    ]
    notes = [
        "Full name",
        "Instructor/Admin/…",
        "Dept/Section",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "1=present, 0=absent",
        "=SUM(D3:I3)",
        "=6-J3",
        "=J3/6*100",
        "Auto-filled",
    ]

    add_header_row(ws, 1, headers, style)
    add_note_row(ws, 2, notes, note_style)

    samples = [
        [
            "Nada Khaled",
            "Instructor",
            "Coding",
            1,
            1,
            1,
            0,
            1,
            1,
            5,
            1,
            83.3,
            "Esraa/Staff Attendance Sheet.xlsx",
        ],
        [
            "Mariam Samir",
            "Instructor",
            "Robotics",
            1,
            0,
            1,
            1,
            1,
            1,
            5,
            1,
            83.3,
            "Esraa/Staff Attendance Sheet.xlsx",
        ],
    ]
    add_sample_rows(ws, 3, samples)
    freeze_and_filter(ws, "A3")

    ws.insert_rows(1)
    title = ws.cell(
        row=1, column=1, value="STAFF ATTENDANCE REGISTER — Standardized Template v1.0"
    )
    title.font = Font(bold=True, size=12, color="843C0C", name="Calibri")
    title.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 22

    path = TEMPLATE_DIR / "StaffAttendance_Template.xlsx"
    wb.save(str(path))
    print(f"  [OK] {path.name}")
    return path


# ── Template 4: CRM / SAT ─────────────────────────────────────────────────────
def create_crm_sat_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM_SAT"

    style = make_header_style(COLORS["header_purple"])
    note_style = make_note_style()

    headers = [
        ("Client Name", 22),
        ("Phone", 16),
        ("Group Code", 14),
        ("Enrollment Status", 18),
        ("Course / SAT Group", 22),
        ("Score / Grade", 14),
        ("Payment Status", 16),
        ("Referral Source", 18),
        ("Notes", 30),
        ("Source File", 28),
    ]
    notes = [
        "Full name",
        "11-digit",
        "e.g. SAT-A1",
        "Active/Completed/Pending",
        "Name",
        "Numeric",
        "Paid/Partial/Unpaid",
        "Social/Referral/…",
        "Any extra info",
        "Auto-filled",
    ]

    add_header_row(ws, 1, headers, style)
    add_note_row(ws, 2, notes, note_style)

    samples = [
        [
            "Mohamed Hassan",
            "01056789012",
            "SAT-B2",
            "Active",
            "SAT Group B",
            None,
            "Paid",
            "Social Media",
            "",
            "CRM/SAT.xlsx",
        ],
    ]
    add_sample_rows(ws, 3, samples)
    freeze_and_filter(ws, "A3")

    dv_status = DataValidation(
        type="list", formula1='"Active,Completed,Pending,Cancelled"', showDropDown=False
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = "D3:D1000"

    ws.insert_rows(1)
    title = ws.cell(
        row=1, column=1, value="CRM & SAT RECORDS — Standardized Template v1.0"
    )
    title.font = Font(bold=True, size=12, color="3B1F75", name="Calibri")
    title.fill = PatternFill(fill_type="solid", fgColor="E8E0F5")
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 22

    path = TEMPLATE_DIR / "CRM_SAT_Template.xlsx"
    wb.save(str(path))
    print(f"  [OK] {path.name}")
    return path


# ── Template 5: Master Unified Flat Table ─────────────────────────────────────
def create_master_schema_template():
    wb = openpyxl.Workbook()

    # Sheet 1: Master flat table
    ws = wb.active
    ws.title = "UnifiedTable"
    style = make_header_style(COLORS["master_header"], font_size=9)
    note_style = make_note_style()

    headers = [
        ("record_id", 14),
        ("source_file", 40),
        ("source_sheet", 18),
        ("domain", 22),
        ("migrated_at", 18),
        ("person_name", 22),
        ("phone", 16),
        ("course_name", 22),
        ("course_level", 12),
        ("session_day", 14),
        ("session_time", 14),
        ("instructor_name", 20),
        ("team_name", 20),
        ("member_list", 35),
        ("competition_category", 20),
        ("group_code", 14),
        ("enrollment_status", 18),
        ("score", 10),
        ("attendance_date", 16),
        ("attendance_status", 16),
        ("source_row_number", 16),
        ("extra_fields", 20),
        ("quality_flags", 25),
    ]
    notes = [
        "UUID (auto)",
        "Relative path to source file",
        "Sheet tab name",
        "CourseAttendance|Competition|StaffAttendance|CRM_SAT|Unclassified",
        "ISO timestamp",
        "Student/staff name",
        "Phone number",
        "Course/subject",
        "Integer level",
        "Day name",
        "HH:MM-HH:MM",
        "Instructor",
        "Team name (competition)",
        "Semicolon-separated members",
        "FLL/Robofest/…",
        "SAT/CRM group code",
        "Active/Completed/Pending",
        "Numeric score",
        "Date of attendance record",
        "1=present, 0=absent, blank=unknown",
        "Row # in source sheet",
        "JSON blob for overflow cols",
        "Pipe-sep QA issues",
    ]

    add_header_row(ws, 1, headers, style, row_height=35)
    add_note_row(ws, 2, notes, note_style, row_height=40)

    # Sample row
    samples = [
        [
            "uuid-0001",
            "courses 2/Html sat 6-8.xlsx",
            "Sheet1",
            "CourseAttendance",
            "2026-02-19T21:00:00",
            "Ahmed Mohamed",
            "01012345678",
            "HTML & CSS",
            1,
            "Saturday",
            "10:00-12:00",
            "Nada",
            None,
            None,
            None,
            None,
            None,
            None,
            "2024-03-02",
            1,
            2,
            None,
            None,
        ],
    ]
    add_sample_rows(ws, 3, samples)
    freeze_and_filter(ws, "A3")

    # Sheet 2: Schema Dictionary
    ws2 = wb.create_sheet("SchemaDictionary")
    dict_headers = [
        ("Field Name", 22),
        ("Data Type", 16),
        ("Nullable", 10),
        ("Valid Values / Format", 36),
        ("Source Columns (examples)", 40),
        ("Notes", 40),
    ]
    dict_style = make_header_style(COLORS["header_teal"])
    add_header_row(ws2, 1, dict_headers, dict_style)

    schema_dict = [
        ["record_id", "UUID/TEXT", "No", "Auto-generated UUID", "—", "Primary key"],
        [
            "source_file",
            "TEXT",
            "No",
            "Relative path string",
            "—",
            "Full relative path from root",
        ],
        ["source_sheet", "TEXT", "Yes", "Sheet tab name", "—", ""],
        [
            "domain",
            "ENUM",
            "No",
            "CourseAttendance|Competition|StaffAttendance|CRM_SAT|Unclassified",
            "—",
            "Assigned by 02_schema_analysis.py",
        ],
        ["migrated_at", "TIMESTAMP", "No", "ISO 8601", "—", "Set at migration time"],
        [
            "person_name",
            "TEXT",
            "Yes",
            "Any string",
            "name, student_name, اسم_الطالب",
            "",
        ],
        [
            "phone",
            "TEXT",
            "Yes",
            "10–15 digits",
            "phone, mobile, رقم",
            "Stored as text to preserve leading zeros",
        ],
        ["course_name", "TEXT", "Yes", "Any string", "course, subject, class", ""],
        ["course_level", "INTEGER", "Yes", "Positive integer", "level, lv, lvl", ""],
        [
            "session_day",
            "TEXT",
            "Yes",
            "Day name (English)",
            "day, weekday",
            "Normalized to full day name",
        ],
        ["session_time", "TEXT", "Yes", "HH:MM-HH:MM", "time, time_slot", ""],
        [
            "instructor_name",
            "TEXT",
            "Yes",
            "Any string",
            "instructor, teacher, coach",
            "",
        ],
        [
            "team_name",
            "TEXT",
            "Yes",
            "Any string",
            "team, team_name, فريق",
            "Competition domain only",
        ],
        [
            "member_list",
            "TEXT",
            "Yes",
            "Name1; Name2; …",
            "members, participants",
            "Semicolon-separated",
        ],
        [
            "competition_category",
            "TEXT",
            "Yes",
            "FLL/Robofest/WRO/…",
            "category, event_type",
            "",
        ],
        [
            "group_code",
            "TEXT",
            "Yes",
            "Any string",
            "group, batch, sat_group",
            "CRM/SAT domain",
        ],
        [
            "enrollment_status",
            "TEXT",
            "Yes",
            "Active|Completed|Pending",
            "status, payment, paid",
            "",
        ],
        ["score", "FLOAT", "Yes", "Any numeric", "score, points, grade", ""],
        [
            "attendance_date",
            "DATE",
            "Yes",
            "ISO 8601 YYYY-MM-DD",
            "Date columns, dd/mm/…",
            "Unpivoted from wide format",
        ],
        [
            "attendance_status",
            "SMALLINT",
            "Yes",
            "0 or 1",
            "Cell values: P,A,1,0,✓,✗,…",
            "1=present, 0=absent",
        ],
        [
            "source_row_number",
            "INTEGER",
            "No",
            "Positive integer",
            "—",
            "Original row in source sheet",
        ],
        [
            "extra_fields",
            "JSON/TEXT",
            "Yes",
            "JSON object",
            "—",
            "Overflow columns not in schema",
        ],
        [
            "quality_flags",
            "TEXT",
            "Yes",
            "pipe-delimited strings",
            "—",
            "e.g. NULL_PHONE|AMBIGUOUS_DATE",
        ],
    ]
    for i, row in enumerate(schema_dict, start=2):
        fill = (
            PatternFill(fill_type="solid", fgColor=COLORS["alt_row"])
            if i % 2 == 0
            else None
        )
        for col_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=i, column=col_idx, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = "A2"

    # Sheet 3: Validation Rules
    ws3 = wb.create_sheet("ValidationRules")
    rv_style = make_header_style(COLORS["header_teal"])
    rv_headers = [
        ("Rule ID", 10),
        ("Description", 40),
        ("SQL / Logic", 60),
        ("Severity", 12),
    ]
    add_header_row(ws3, 1, rv_headers, rv_style)
    rules = [
        [
            "RC-01",
            "domain must be one of 5 valid values",
            "domain IN ('CourseAttendance','Competition','StaffAttendance','CRM_SAT','Unclassified')",
            "ERROR",
        ],
        [
            "RC-02",
            "attendance_status must be 0, 1, or NULL",
            "attendance_status IN (0,1) OR attendance_status IS NULL",
            "ERROR",
        ],
        ["RC-03", "source_file must not be NULL", "source_file IS NOT NULL", "ERROR"],
        [
            "RC-04",
            "migrated_at must be after 2020-01-01",
            "migrated_at > '2020-01-01'",
            "ERROR",
        ],
        [
            "RC-05",
            "If attendance_date set, person_name must be set",
            "NOT (attendance_date IS NOT NULL AND person_name IS NULL)",
            "WARNING",
        ],
        [
            "RC-06",
            "phone format: 10-15 digits only",
            "phone REGEXP '^[0-9]{10,15}$' OR phone IS NULL",
            "WARNING",
        ],
        [
            "RC-07",
            "course_level must be positive integer if present",
            "course_level > 0 OR course_level IS NULL",
            "WARNING",
        ],
        [
            "RC-08",
            "No duplicate (source_file, source_sheet, row, date)",
            "UNIQUE(source_file, source_sheet, source_row_number, attendance_date)",
            "ERROR",
        ],
        [
            "SEM-01",
            "Attendance date not in the future",
            "attendance_date <= CURRENT_DATE OR attendance_date IS NULL",
            "WARNING",
        ],
        [
            "SEM-02",
            "Attendance rate 10–100% per course",
            "AVG(attendance_status) BETWEEN 0.10 AND 1.00 GROUP BY course_name",
            "INFO",
        ],
    ]
    for i, rule in enumerate(rules, start=2):
        for col_idx, val in enumerate(rule, start=1):
            cell = ws3.cell(row=i, column=col_idx, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if rule[3] == "ERROR":
                cell.fill = PatternFill(fill_type="solid", fgColor="FFD7D7")
            elif rule[3] == "WARNING":
                cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        ws3.row_dimensions[i].height = 28
    ws3.freeze_panes = "A2"
    for col_letter, width in zip(["A", "B", "C", "D"], [10, 40, 60, 12]):
        ws3.column_dimensions[col_letter].width = width

    path = TEMPLATE_DIR / "MasterSchema_UnifiedTable.xlsx"
    wb.save(str(path))
    print(f"  [OK] {path.name}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def generate_all_templates():
    print(f"\n{'=' * 70}")
    print(f"  Schema Template Generator")
    print(f"  Output directory: {TEMPLATE_DIR}")
    print(f"{'=' * 70}\n")

    paths = []
    paths.append(create_course_attendance_template())
    paths.append(create_competition_template())
    paths.append(create_staff_attendance_template())
    paths.append(create_crm_sat_template())
    paths.append(create_master_schema_template())

    print(f"\n{'=' * 70}")
    print(f"  {len(paths)} templates generated successfully.")
    print(f"  Location: {TEMPLATE_DIR}")
    print(f"{'=' * 70}\n")
    return paths


if __name__ == "__main__":
    generate_all_templates()
