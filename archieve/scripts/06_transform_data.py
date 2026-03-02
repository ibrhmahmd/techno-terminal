import os
import sys
import re
import json
import uuid
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl

# Force UTF-8 for Windows terminal printing
sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "data" / "processed" / "_standardization_output"
REORG_DIR = BASE_DIR / "data" / "processed" / "_REORGANIZED"
MASTER_DB_FILE = OUTPUT_DIR / "Master_Database.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# MAPPING DICTIONARIES (Extending from 02_schema_analysis.py)
# ─────────────────────────────────────────────────────────────────────────────
FIELD_MAP = {
    # Person Name
    "name": "person_name",
    "student_name": "person_name",
    "student": "person_name",
    "pupils": "person_name",
    "child_name": "person_name",
    "اسم_الطالب": "person_name",
    "full_name": "person_name",
    "name_full": "person_name",
    "الاسم": "person_name",
    "childname": "person_name",
    "studentname": "person_name",
    "employee": "person_name",
    "employee_name": "person_name",
    "staff": "person_name",
    "staff_name": "person_name",
    "client": "person_name",
    "client_name": "person_name",
    "customer": "person_name",
    # Phone
    "phone": "phone",
    "mobile": "phone",
    "tel": "phone",
    "telephone": "phone",
    "contact": "phone",
    "phone_number": "phone",
    "mobile_number": "phone",
    "رقم_التليفون": "phone",
    "رقم_الهاتف": "phone",
    "رقم": "phone",
    "number": "phone",
    # Course Name
    "course": "course_name",
    "subject": "course_name",
    "class": "course_name",
    "program": "course_name",
    "course_name": "course_name",
    # Level
    "level": "course_level",
    "lv": "course_level",
    "lvl": "course_level",
    "course_level": "course_level",
    "grade": "course_level",
    # Day & Time
    "day": "session_day",
    "session_day": "session_day",
    "weekday": "session_day",
    "time": "session_time",
    "time_slot": "session_time",
    "from_to": "session_time",
    "start_time": "session_time",
    "end_time": "session_time",
    "start": "session_time",
    # Instructor
    "instructor": "instructor_name",
    "teacher": "instructor_name",
    "coach": "instructor_name",
    "trainer": "instructor_name",
    # Team (Competitions)
    "team": "team_name",
    "team_name": "team_name",
    "فريق": "team_name",
    "اسم_الفريق": "team_name",
    "robot": "team_name",
    "robot_name": "team_name",
    "members": "member_list",
    "member_list": "member_list",
    "participants": "member_list",
    # Competition Category
    "category": "competition_category",
    "event": "competition_category",
    "event_type": "competition_category",
    "fll": "competition_category",
    "robofest": "competition_category",
    "wro": "competition_category",
    "sumo": "competition_category",
    "parade": "competition_category",
    "arts": "competition_category",
    "competition": "competition_category",
    # Scores & Status
    "score": "score",
    "points": "score",
    "result": "score",
    "final_score": "score",
    "mark": "score",
    "status": "enrollment_status",
    "enrollment": "enrollment_status",
    # CRM/Groups
    "sat": "group_code",
    "sat_group": "group_code",
    "group": "group_code",
    "group_code": "group_code",
    "batch": "group_code",
    "batch_code": "group_code",
    # Payments & Age
    "payment": "payment_status",
    "paid": "payment_status",
    "fee_paid": "payment_status",
    "month_paid": "payment_status",
    "fees": "payment_status",
    "age": "age",
    "birth_date": "birth_date",
}

DATE_COLUMN_PATTERN = re.compile(
    r"^\d{1,2}[_/\-\.]\d{1,2}([_/\-\.]\d{2,4})?$|"
    r"^(sat|sun|mon|tue|wed|thu|fri)\d*$|"
    r"^week\d+$|^session\d+$|^s_?\d+$|^w_?\d+$|^d_?\d+$",
    re.IGNORECASE,
)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def normalize_header(header_str):
    if not isinstance(header_str, str):
        return ""
    # lower, strip, replace spaces/dots with underscores
    val = str(header_str).lower().strip()
    val = re.sub(r"[\s\.\-]+", "_", val)
    return val


def identify_header_row(sheet):
    """Finds the row index (0-based) that likely contains the headers."""
    best_row_idx = 0
    max_strings = 0

    # Check first 10 rows
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        string_cnt = sum(
            1 for cell in row if isinstance(cell, str) and len(cell.strip()) > 0
        )
        if string_cnt > max_strings:
            max_strings = string_cnt
            best_row_idx = i

    return best_row_idx


def map_column_name(raw_header):
    norm = normalize_header(raw_header)
    if not norm:
        return "unknown", False

    # 1. Exact match in synonym map
    if norm in FIELD_MAP:
        return FIELD_MAP[norm], False

    # 2. Check if it's a date/attendance column
    if DATE_COLUMN_PATTERN.match(norm) or "date" in norm:
        return norm, True  # True means it's an attendance parameter to be unpivoted

    # 3. Otherwise, it's an extra field
    return norm, False


def parse_attendance_status(val):
    """Standardize attendance markings into boolean 1/0/None"""
    if val is None or str(val).strip() == "":
        return None
    val_str = str(val).lower().strip()
    # Positive markers
    if val_str in [
        "1",
        "true",
        "yes",
        "y",
        "p",
        "present",
        "حضر",
        "✓",
        "v",
        "ok",
        "attend",
    ]:
        return 1
    # Negative markers
    if val_str in ["0", "false", "no", "n", "a", "absent", "غاب", "x", "late"]:
        return 0
    # If it's a date or string, they might be marking time of arrival, so count as present
    return 1


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ETL PROCESS
# ─────────────────────────────────────────────────────────────────────────────
def transform_data():
    if not REORG_DIR.exists():
        print(f"[ERROR] Reorganized directory {REORG_DIR} not found.")
        return

    print(f"\n{'=' * 70}")
    print(f"  Starting Phase 2: Data Extraction & Schema Mapping")
    print(f"{'=' * 70}\n")

    master_records = []
    files_processed = 0
    errors = 0

    # Walk through _REORGANIZED folder
    for root, dirs, files in os.walk(REORG_DIR):
        for file in files:
            if not file.endswith((".xlsx", ".xlsm", ".xls")) or file.startswith("~$"):
                continue

            file_path = Path(root) / file
            rel_path = file_path.relative_to(REORG_DIR)

            # Determine Domain from folder structure
            if "01_CourseAttendance" in root:
                domain = "CourseAttendance"
            elif "02_CompetitionTracking" in root:
                domain = "CompetitionTracking"
            elif "03_StaffAttendance" in root:
                domain = "StaffAttendance"
            elif "04_CRM_SAT" in root:
                domain = "CRM_SAT"
            else:
                domain = "Unclassified"

            if domain == "Unclassified":
                continue  # Skip corrupt/unclassified files for the master db

            print(f"  - Extracting: {rel_path}")

            try:
                # Use openpyxl with data_only=True to get values, not formulas
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    header_row_idx = identify_header_row(sheet)

                    rows_iter = sheet.iter_rows(values_only=True)
                    rows = list(rows_iter)

                    if len(rows) <= header_row_idx + 1:
                        continue  # Empty sheet

                    raw_headers = rows[header_row_idx]
                    data_rows = rows[header_row_idx + 1 :]

                    # Map headers
                    col_mapping = []  # list of (idx, mapped_name, is_attendance_col)
                    for col_idx, h in enumerate(raw_headers):
                        if h is not None and str(h).strip() != "":
                            mapped_name, is_att = map_column_name(h)
                            col_mapping.append(
                                (col_idx, mapped_name, is_att, str(h).strip())
                            )

                    if not col_mapping:
                        continue  # No valid headers

                    # Pre-scan the next 5-10 rows to look for metadata (Session Date, Teacher, etc)
                    # and find the actual dates if the headers are just "s1", "s2" etc.
                    actual_date_mapping = {}  # maps col_idx -> real date string
                    first_student_row_idx = header_row_idx + 1

                    for r_idx, r_data in enumerate(data_rows):
                        first_cell = (
                            str(r_data[0]).lower().strip()
                            if r_data and r_data[0]
                            else ""
                        )
                        name_cell = ""
                        # Try to find the person_name mapped column to check for metadata
                        for c_idx, m_name, _, _ in col_mapping:
                            if m_name == "person_name" and c_idx < len(r_data):
                                name_cell = str(r_data[c_idx]).lower().strip()
                                break

                        check_str = first_cell if first_cell else name_cell

                        # Stop if it looks like a real student
                        if not any(
                            k in check_str
                            for k in [
                                "date",
                                "time",
                                "today",
                                "oday",
                                "session",
                                "teacher",
                                "instructor",
                                "notes",
                                "none",
                                "nan",
                                "course",
                            ]
                        ):
                            if check_str != "" and check_str != "nan":
                                first_student_row_idx = header_row_idx + 1 + r_idx
                                break
                        else:
                            # It is a metadata row. Check if it's the "Session Date" row
                            if "date" in check_str or "تاريخ" in check_str:
                                for c_idx, m_name, is_att, orig_header in col_mapping:
                                    if is_att and c_idx < len(r_data):
                                        val = r_data[c_idx]
                                        if (
                                            val is not None
                                            and str(val).strip() != ""
                                            and str(val).strip().lower() != "nan"
                                        ):
                                            # Found an actual date for this column
                                            actual_date_mapping[c_idx] = (
                                                str(val).split()[0].strip()
                                            )  # just the date part if it's a datetime

                    # Process student rows
                    for row_idx, row_data in enumerate(
                        data_rows[first_student_row_idx - (header_row_idx + 1) :],
                        start=first_student_row_idx + 1,
                    ):
                        # Skip completely empty rows
                        if all(
                            val is None or str(val).strip() == "" for val in row_data
                        ):
                            continue

                        # Double check we aren't pulling in a straggler metadata row
                        name_val = ""
                        for c_idx, m_name, _, _ in col_mapping:
                            if m_name == "person_name" and c_idx < len(row_data):
                                name_val = str(row_data[c_idx]).lower().strip()
                                break
                        if any(
                            k in name_val
                            for k in [
                                "date",
                                "time",
                                "today",
                                "oday",
                                "session",
                                "teacher",
                                "instructor",
                                "none",
                                "nan",
                                "course",
                            ]
                        ):
                            continue

                        # Base record container
                        base_rec = {
                            "record_id": str(uuid.uuid4()),
                            "source_file": str(rel_path),
                            "source_sheet": sheet_name,
                            "domain": domain,
                            "migrated_at": datetime.now().isoformat(),
                            "source_row_number": row_idx,
                            "person_name": None,
                            "phone": None,
                            "course_name": None,
                            "course_level": None,
                            "session_day": None,
                            "session_time": None,
                            "instructor_name": None,
                            "team_name": None,
                            "member_list": None,
                            "competition_category": None,
                            "group_code": None,
                            "enrollment_status": None,
                            "score": None,
                            "payment_status": None,
                            "age": None,
                            "birth_date": None,
                            "attendance_date": None,
                            "attendance_status": None,
                            "extra_fields": {},
                        }

                        attendance_dates_data = []  # List of (date_str, status_val)

                        for col_idx, mapped_name, is_att, orig_header in col_mapping:
                            if col_idx >= len(row_data):
                                continue
                            cell_val = row_data[col_idx]

                            if cell_val is None or str(cell_val).strip() == "":
                                continue

                            if is_att:
                                # This is a wide-format date column, store for unpivoting
                                att_status = parse_attendance_status(cell_val)
                                if att_status is not None:
                                    attendance_dates_data.append(
                                        (orig_header, att_status)
                                    )
                            elif mapped_name in base_rec:
                                # Standard mapped column
                                base_rec[mapped_name] = str(cell_val).strip()
                            else:
                                # Extra unmapped column goes into JSON/dict
                                base_rec["extra_fields"][orig_header] = str(
                                    cell_val
                                ).strip()

                        # Only serialize extra fields at the end
                        base_rec["extra_fields"] = json.dumps(
                            base_rec["extra_fields"], ensure_ascii=False
                        )

                        # Unpivot step: if we found attendance dates, emit one row PER DATE
                        if attendance_dates_data:
                            for att_date, att_status in attendance_dates_data:
                                new_rec = base_rec.copy()
                                new_rec["record_id"] = str(
                                    uuid.uuid4()
                                )  # Unique ID per unpivoted row
                                new_rec["attendance_date"] = str(att_date)
                                new_rec["attendance_status"] = att_status
                                master_records.append(new_rec)
                        else:
                            # No attendance grid data, just emit the base record
                            master_records.append(base_rec)

                wb.close()
                files_processed += 1

            except Exception as e:
                print(f"  [ERROR] Failed to extract from {rel_path} - {e}")
                errors += 1

    # ── Write output DataFrame ───────────────────────────────────────────────
    if not master_records:
        print("\n[WARNING] No data extracted across all files.")
        return

    df = pd.DataFrame(master_records)

    # Optional: Reorder columns for readability
    cols = [
        "record_id",
        "source_file",
        "source_sheet",
        "domain",
        "person_name",
        "phone",
        "course_name",
        "course_level",
        "session_day",
        "session_time",
        "instructor_name",
        "attendance_date",
        "attendance_status",
        "team_name",
        "member_list",
        "competition_category",
        "group_code",
        "score",
        "payment_status",
        "age",
        "birth_date",
        "enrollment_status",
        "source_row_number",
        "migrated_at",
        "extra_fields",
    ]

    # Only keep columns that exist (in case typo)
    cols = [c for c in cols if c in df.columns]
    df = df[cols]

    print(f"\n{'=' * 70}")
    print(f"  Extraction Complete")
    print(f"  Files processesd: {files_processed}")
    print(f"  Total records generated: {len(df)} rows")
    if errors > 0:
        print(f"  Errors encountered: {errors}")
    print(f"{'=' * 70}\n")

    print(f"  Saving to {MASTER_DB_FILE}...")
    df.to_excel(MASTER_DB_FILE, index=False, engine="openpyxl")
    print(f"  [OK] Saved successfully!")


if __name__ == "__main__":
    try:
        transform_data()
    except Exception as e:
        import traceback

        traceback.print_exc()
